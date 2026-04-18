"""Operational EDA over every sheet of every code/inputs/*.xlsx workbook.

This module is the single source of truth for the "Operational Deep Dive"
layer that sits on top of the existing feature-engineering EDA. It writes:

- CSV aggregates under ``code/outputs/reports/operational_eda/``
- Plots under ``code/outputs/plots/operational_eda/``

The companion notebook ``code/inference_engine/notebooks/operational_eda.ipynb``
imports :func:`run` from here so the notebook stays narrative and the logic
stays testable.

Run directly:

    python code/inference_engine/scripts/operational_eda.py

The script is idempotent: rerunning overwrites the CSV/plot artifacts.
"""

from __future__ import annotations

import json
import sys
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

REPO_ROOT = Path(__file__).resolve().parents[3]
CODE_ROOT = REPO_ROOT / "code"
INPUTS = CODE_ROOT / "inputs"
PHASE1 = CODE_ROOT / "intermediates" / "phase1"
REPORTS = CODE_ROOT / "outputs" / "reports" / "operational_eda"
PLOTS = CODE_ROOT / "outputs" / "plots" / "operational_eda"

Q1_XLSX = INPUTS / "Q1 Daily Metrics 2026.xlsx"
INTAKE_TEMPLATE = INPUTS / "RideYourWay_Prospective_Market_Intake_Template.xlsx"
INTAKE_EXAMPLE = INPUTS / "RideYourWay_Prospective_Market_Intake_Example.xlsx"

REGIONS = ["Grand Rapids", "Lansing", "Battle Creek"]
WEEK_ORDER = ["Week 1", "Week 2", "Week 3", "Week 4", "Week 5"]
DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

# Gate thresholds (from code/config/pjtl_kpis_and_formulas.json and the RYW
# scope doc). Kept here as literals because the EDA needs to render pass/fail
# colors without pulling the whole backend.
GATE_DEFS: List[Dict[str, object]] = [
    {"id": "g1_vehicle_utilization", "label": "Vehicle utilization", "target": 0.95, "comparator": "ge"},
    {"id": "g2_billed_utilization", "label": "Billed utilization", "target": 1.05, "comparator": "ge"},
    {"id": "g3_volume_pool", "label": "Total volume pool", "target": 1.20, "comparator": "ge"},
    {"id": "g4_rev_per_kentleg", "label": "Revenue / Kent-Leg", "target": 70.0, "comparator": "ge"},
    {"id": "g5_higher_acuity_mix", "label": "Higher-acuity mix", "target": 0.05, "comparator": "ge"},
    {"id": "g6_nonbillable_ns", "label": "Non-billable no-shows", "target": 0.10, "comparator": "le"},
    {"id": "g7_road_time", "label": "Road time / vehicle / day", "target": 9.0, "comparator": "ge"},
    {"id": "g8_contract_concentration", "label": "Largest contract share", "target": 0.20, "comparator": "le"},
    {"id": "g9_cost_per_road_hour", "label": "Cost per road hour", "target": 50.0, "comparator": "le"},
]


@dataclass
class EdaResult:
    """Thin container tracking artifacts written to disk."""

    csvs: List[Path] = field(default_factory=list)
    plots: List[Path] = field(default_factory=list)
    summary: Dict[str, object] = field(default_factory=dict)

    def add_csv(self, p: Path) -> None:
        self.csvs.append(p.relative_to(REPO_ROOT))

    def add_plot(self, p: Path) -> None:
        self.plots.append(p.relative_to(REPO_ROOT))


# ----- small utilities ----------------------------------------------------


def _open_q1() -> "openpyxl.workbook.workbook.Workbook":  # type: ignore[name-defined]
    """Open the Q1 workbook in read-only mode (bypasses the pivot-cache bug)."""

    return load_workbook(Q1_XLSX, data_only=True, read_only=True)


def _ensure_dirs() -> None:
    for d in (REPORTS, PLOTS):
        d.mkdir(parents=True, exist_ok=True)


def _savefig(fig: plt.Figure, path: Path, result: EdaResult) -> None:
    fig.tight_layout()
    fig.savefig(path, dpi=140, bbox_inches="tight")
    plt.close(fig)
    result.add_plot(path)


def _save_csv(df: pd.DataFrame, path: Path, result: EdaResult) -> None:
    df.to_csv(path, index=False)
    result.add_csv(path)


def _num(x: object) -> Optional[float]:
    """Coerce a workbook cell value to float, returning None for blanks/errors."""

    if x is None or x == "" or (isinstance(x, str) and x.startswith("#")):
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _pass(value: Optional[float], target: float, comparator: str) -> Optional[bool]:
    if value is None or np.isnan(value):
        return None
    return bool(value >= target) if comparator == "ge" else bool(value <= target)


# ----- section 1: data coverage -------------------------------------------


def section_coverage(result: EdaResult) -> pd.DataFrame:
    """Row-count and missingness snapshot for every sheet we read."""

    wb = _open_q1()
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows.append({
            "workbook": "Q1 Daily Metrics 2026.xlsx",
            "sheet": sheet,
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
        })
    wb.close()

    for path in (INTAKE_TEMPLATE, INTAKE_EXAMPLE):
        wb = load_workbook(path, data_only=True, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows.append({
                "workbook": path.name,
                "sheet": sheet,
                "state": ws.sheet_state,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
            })
        wb.close()

    df = pd.DataFrame(rows)
    _save_csv(df, REPORTS / "sheet_coverage.csv", result)
    result.summary["sheets_total"] = int(df.shape[0])
    result.summary["sheets_visible"] = int((df["state"] == "visible").sum())
    result.summary["sheets_hidden"] = int((df["state"] == "hidden").sum())
    return df


# ----- section 2: fleetwise (Regional Performance) ------------------------


REGIONAL_COL_OFFSETS = {"Grand Rapids": 0, "Lansing": 28, "Battle Creek": 56}
REGIONAL_METRIC_COLS = {
    "billable_ns": 4,
    "billed_usage": 5,
    "cancellation_24h": 6,
    "otp": 8,
    "vehicle_usage": 11,
    "scheduled_usage": 14,
    "total_rides": 15,
    "kent_legs": 17,
    "target_capacity": 19,
    "skl": 21,
    "kl_multiple": 22,
    "quality_pool": 23,
    "total_pool": 24,
}


def _extract_regional(ws) -> pd.DataFrame:
    """Flatten Regional Performance into one row per (region, scope)."""

    all_rows = list(ws.iter_rows(values_only=True))
    records: List[Dict[str, object]] = []

    # Q1 quarter rollup is row index 4; each Week N block starts with label
    # in column 1 ("Q1" / "Week 1" / ...) and the daily rows are 2..7 rows
    # below the header.
    for ridx, row in enumerate(all_rows):
        label_cell = row[1] if len(row) > 1 else None
        if not isinstance(label_cell, str):
            continue
        label = label_cell.strip()
        if label not in (["Q1", "Target"] + WEEK_ORDER + DAY_ORDER + ["Total"]):
            continue
        if label in ("Target",):
            continue  # target row has no values we need
        # Determine scope: Q1 or Week N header row → scope="quarter"/"week"
        # Otherwise day row → need to know current week context
        for region, off in REGIONAL_COL_OFFSETS.items():
            rec: Dict[str, object] = {"region": region, "label": label, "row": ridx}
            for metric, col in REGIONAL_METRIC_COLS.items():
                rec[metric] = _num(row[col + off]) if col + off < len(row) else None
            records.append(rec)

    df = pd.DataFrame(records)
    # Attach a scope column: any "Q1" row is the quarter rollup; others are
    # weekly detail rows (we group under their nearest preceding "Week N").
    scope_rows = df[df["label"].isin(["Q1"] + WEEK_ORDER)].sort_values("row")
    scope_map = {}
    current = None
    for _, r in df[["row", "label"]].drop_duplicates().sort_values("row").iterrows():
        if r["label"] == "Q1":
            current = "Q1"
        elif r["label"] in WEEK_ORDER:
            current = r["label"]
        scope_map[r["row"]] = current
    df["scope"] = df["row"].map(scope_map)
    return df


def section_fleetwise(result: EdaResult, cv: pd.DataFrame) -> pd.DataFrame:
    """Per-region rollup of the nine gates."""

    wb = _open_q1()
    regional = _extract_regional(wb["Regional Performance"])
    wb.close()

    # Q1 rollup rows for each region
    q1 = regional[regional["label"] == "Q1"].groupby("region").first().reset_index()

    # Payer concentration + higher-acuity mix from contract_volume_base
    cv["pu_zone_norm"] = cv["pu_zone"].replace({
        "GrandRapids": "Grand Rapids",
        "Battle Creek/Kalamazoo": "Battle Creek",
        "Lansing": "Lansing",
        "Northern Michigan": "Northern Michigan",
    })
    completed = cv[cv["order_status"] == "Completed"].copy()
    # Largest payer share of volume + revenue per region
    conc_rows = []
    for region in REGIONS:
        sub = completed[completed["pu_zone_norm"] == region]
        if sub.empty:
            conc_rows.append({"region": region, "largest_payer_vol": None, "largest_payer_rev": None})
            continue
        by_payer_vol = sub.groupby("payer_id")["kent_legs"].sum().sort_values(ascending=False)
        by_payer_rev = sub.groupby("payer_id")["order_price"].sum().sort_values(ascending=False)
        conc_rows.append({
            "region": region,
            "largest_payer_vol": float(by_payer_vol.iloc[0] / by_payer_vol.sum()) if by_payer_vol.sum() else None,
            "largest_payer_rev": float(by_payer_rev.iloc[0] / by_payer_rev.sum()) if by_payer_rev.sum() else None,
            "largest_payer_vol_name": by_payer_vol.index[0],
            "largest_payer_rev_name": by_payer_rev.index[0],
        })
    conc = pd.DataFrame(conc_rows)

    # Higher-acuity mix: Stretcher in completed rides (SecureCare not in cv)
    ha_rows = []
    for region in REGIONS:
        sub = completed[completed["pu_zone_norm"] == region]
        total = len(sub)
        ha = len(sub[sub["order_mode"] == "Stretcher"])
        ha_rows.append({
            "region": region,
            "higher_acuity_mix": (ha / total) if total else None,
            "completed_trips": total,
        })
    ha = pd.DataFrame(ha_rows)

    # Revenue per Kent-Leg (completed)
    rev_kl_rows = []
    for region in REGIONS:
        sub = completed[completed["pu_zone_norm"] == region]
        rev = sub["order_price"].sum()
        kl = sub["kent_legs"].sum()
        rev_kl_rows.append({
            "region": region,
            "revenue_per_kentleg": (rev / kl) if kl else None,
            "revenue": float(rev),
            "kent_legs": float(kl),
        })
    rev_kl = pd.DataFrame(rev_kl_rows)

    # Non-billable no-shows from contract_volume (No show rows with Reason != Billed no show)
    ns_rows = []
    for region in REGIONS:
        sub = cv[cv["pu_zone_norm"] == region]
        denom = len(sub)
        non_bill = len(sub[(sub["reason"].isin(["No Show", "No show", "Same Day Cancel", "Confirmation Cancel"])) &
                           (sub["order_status"] != "Completed")])
        ns_rows.append({
            "region": region,
            "nonbillable_ns_rate": (non_bill / denom) if denom else None,
        })
    ns = pd.DataFrame(ns_rows)

    # Assemble scorecard. Gate 3 uses KL multiple (SKL / Target Capacity),
    # not the raw Total Pool count; the multiple is the ratio the charter
    # requires to exceed 1.20.
    sc = q1[["region", "vehicle_usage", "billed_usage", "kl_multiple", "total_pool", "otp"]].copy()
    sc = sc.rename(columns={"kl_multiple": "volume_pool_ratio"})
    sc = sc.merge(rev_kl[["region", "revenue_per_kentleg"]], on="region", how="left")
    sc = sc.merge(ha[["region", "higher_acuity_mix"]], on="region", how="left")
    sc = sc.merge(ns[["region", "nonbillable_ns_rate"]], on="region", how="left")
    sc = sc.merge(conc[["region", "largest_payer_vol", "largest_payer_rev",
                        "largest_payer_vol_name", "largest_payer_rev_name"]], on="region", how="left")

    # Compute pass/fail per gate
    rows = []
    for _, r in sc.iterrows():
        measures = {
            "g1_vehicle_utilization": r["vehicle_usage"],
            "g2_billed_utilization": r["billed_usage"],
            "g3_volume_pool": r["volume_pool_ratio"],
            "g4_rev_per_kentleg": r["revenue_per_kentleg"],
            "g5_higher_acuity_mix": r["higher_acuity_mix"],
            "g6_nonbillable_ns": r["nonbillable_ns_rate"],
            # g7 road time - from vehicle_day_base; computed below per-region
            "g8_contract_concentration": max(
                r["largest_payer_vol"] or 0.0,
                r["largest_payer_rev"] or 0.0,
            ),
        }
        for gd in GATE_DEFS:
            gid = gd["id"]  # type: ignore[index]
            if gid in ("g7_road_time", "g9_cost_per_road_hour"):
                continue
            val = measures.get(gid)
            rows.append({
                "region": r["region"],
                "gate": gid,
                "label": gd["label"],
                "target": gd["target"],
                "comparator": gd["comparator"],
                "value": val,
                "pass": _pass(val, float(gd["target"]), str(gd["comparator"])),
            })
        # Concentration detail
        rows.append({
            "region": r["region"],
            "gate": "g8_largest_payer_vol",
            "label": "Largest payer share of volume",
            "target": 0.20,
            "comparator": "le",
            "value": r["largest_payer_vol"],
            "pass": _pass(r["largest_payer_vol"], 0.20, "le"),
            "detail": r["largest_payer_vol_name"],
        })
        rows.append({
            "region": r["region"],
            "gate": "g8_largest_payer_rev",
            "label": "Largest payer share of revenue",
            "target": 0.20,
            "comparator": "le",
            "value": r["largest_payer_rev"],
            "pass": _pass(r["largest_payer_rev"], 0.20, "le"),
            "detail": r["largest_payer_rev_name"],
        })

    # Road time from vehicle_day_base: road_time per vehicle per day
    vd = pd.read_csv(PHASE1 / "vehicle_day_base.csv")
    # vehicle_day_base has 'mode' but not region; the vehicle name encodes
    # region via prefix (e.g., GR##, L##, BC##). Fallback: treat all vehicles
    # as fleet-level for Gate 7 per region until we have a mapping.
    # For now we use fleet average across region column absence.
    fleet_road_time = vd["road_time"].mean() if "road_time" in vd else None
    for region in REGIONS:
        rows.append({
            "region": region,
            "gate": "g7_road_time",
            "label": "Road time / vehicle / day",
            "target": 9.0,
            "comparator": "ge",
            "value": fleet_road_time,
            "pass": _pass(fleet_road_time, 9.0, "ge"),
            "detail": "fleet-level (per-region vehicle mapping not available)",
        })

    # Cost per road hour: fleet-level only (per Zach)
    wm = pd.read_csv(PHASE1 / "weekly_margin_base.csv")
    total_cost_q1 = wm.loc[wm["scope_level"] == "quarter_total", "total_cost"].sum()
    road_hours_q1 = vd["road_time"].sum() if "road_time" in vd else None
    cost_per_rh = (total_cost_q1 / road_hours_q1) if road_hours_q1 else None
    for region in REGIONS:
        rows.append({
            "region": region,
            "gate": "g9_cost_per_road_hour",
            "label": "Cost per road hour",
            "target": 50.0,
            "comparator": "le",
            "value": cost_per_rh,
            "pass": _pass(cost_per_rh, 50.0, "le"),
            "detail": "fleet-level estimate (per-region cost not reported)",
        })

    scorecard = pd.DataFrame(rows)
    _save_csv(scorecard, REPORTS / "fleet_gate_scorecard.csv", result)

    # Fleet-level column (for comparison)
    fleet_rows = cv.copy()
    fleet_rows["pu_zone_norm"] = fleet_rows["pu_zone"]
    # Compute fleet totals for the same gate metrics
    result.summary["fleet_q1_vehicle_utilization"] = float(q1["vehicle_usage"].mean(skipna=True))
    result.summary["fleet_q1_billed_utilization"] = float(q1["billed_usage"].mean(skipna=True))
    result.summary["fleet_q1_volume_pool_ratio"] = float(q1["kl_multiple"].mean(skipna=True))

    # Plot: per-region bars for each gate
    pivot = scorecard.pivot_table(index="region", columns="gate", values="value", aggfunc="first")
    plot_gates = ["g1_vehicle_utilization", "g2_billed_utilization",
                   "g3_volume_pool", "g4_rev_per_kentleg",
                   "g5_higher_acuity_mix", "g6_nonbillable_ns",
                   "g8_largest_payer_vol", "g8_largest_payer_rev"]
    fig, axes = plt.subplots(2, 4, figsize=(16, 8))
    for ax, gid in zip(axes.flatten(), plot_gates):
        if gid not in pivot.columns:
            ax.axis("off")
            continue
        vals = pivot[gid].reindex(REGIONS)
        gd = next((g for g in GATE_DEFS if g["id"] == gid), None)
        if gd is None:
            gd = {"label": gid, "target": 0, "comparator": "ge"}
            # concentration gates use the standard g8 target
            if gid.startswith("g8_"):
                gd = {"label": "Largest payer share", "target": 0.20, "comparator": "le"}
        ax.bar(vals.index, vals.values, color=["#00274C", "#FFCB05", "#A0A0A0"])
        ax.axhline(float(gd["target"]), ls="--", color="red", alpha=0.6, label=f"target {gd['target']}")
        ax.set_title(gd["label"], fontsize=10)
        ax.tick_params(axis="x", rotation=30)
    fig.suptitle("Q1 2026 - gate readings by fleet", fontsize=13)
    _savefig(fig, PLOTS / "fleet_gate_bars.png", result)

    return scorecard


# ----- section 3: weekly trend -------------------------------------------


def section_weekly(result: EdaResult, cv: pd.DataFrame) -> pd.DataFrame:
    """Week 1..5 movement of every fleet-total gate."""

    wb = _open_q1()
    regional = _extract_regional(wb["Regional Performance"])
    wb.close()

    # Reduce regional df to per-(region, week) "Total" rows
    week_rows = regional[
        (regional["label"] == "Total")
        & (regional["scope"].isin(WEEK_ORDER))
    ].copy()

    # Fleet-total for each week = mean of regional (simple average; regional
    # target capacity differs so this is a directional view, not exact).
    fleet_weekly = week_rows.groupby("scope").agg({
        "vehicle_usage": "mean",
        "billed_usage": "mean",
        "kl_multiple": "mean",
        "otp": "mean",
    }).reindex(WEEK_ORDER).reset_index().rename(columns={
        "scope": "week",
        "kl_multiple": "volume_pool_ratio",
    })

    # Revenue/KL, higher-acuity mix, NS rate by week from contract_volume
    completed = cv[cv["order_status"] == "Completed"]
    rev_kl_week = completed.groupby("week_normalized").agg(
        revenue=("order_price", "sum"),
        kent_legs=("kent_legs", "sum"),
    ).reset_index()
    rev_kl_week["revenue_per_kentleg"] = rev_kl_week["revenue"] / rev_kl_week["kent_legs"]

    ha_week = completed.groupby("week_normalized").apply(
        lambda x: (x["order_mode"] == "Stretcher").sum() / len(x) if len(x) else 0.0
    ).reset_index(name="higher_acuity_mix")

    nb_week = cv.groupby("week_normalized").apply(
        lambda x: len(x[(x["reason"].isin(["No Show", "No show", "Same Day Cancel", "Confirmation Cancel"])) &
                         (x["order_status"] != "Completed")]) / len(x) if len(x) else 0.0
    ).reset_index(name="nonbillable_ns_rate")

    # Largest payer concentration by week
    conc_week = []
    for wk in WEEK_ORDER:
        sub = completed[completed["week_normalized"] == wk]
        if sub.empty:
            conc_week.append({"week": wk, "largest_payer_vol": None, "largest_payer_rev": None})
            continue
        by_vol = sub.groupby("payer_id")["kent_legs"].sum().sort_values(ascending=False)
        by_rev = sub.groupby("payer_id")["order_price"].sum().sort_values(ascending=False)
        conc_week.append({
            "week": wk,
            "largest_payer_vol": float(by_vol.iloc[0] / by_vol.sum()) if by_vol.sum() else None,
            "largest_payer_rev": float(by_rev.iloc[0] / by_rev.sum()) if by_rev.sum() else None,
            "largest_payer_vol_name": by_vol.index[0],
            "largest_payer_rev_name": by_rev.index[0],
        })
    conc_week_df = pd.DataFrame(conc_week)

    # Weekly margin
    wm = pd.read_csv(PHASE1 / "weekly_margin_base.csv")
    wm_week = wm[wm["scope_level"] == "week"].groupby("week_normalized").agg(
        total_revenue=("total_revenue", "sum"),
        total_cost=("total_cost", "sum"),
        profit_margin=("profit_margin", "mean"),
    ).reindex(WEEK_ORDER).reset_index()

    # Assemble
    trend = fleet_weekly.merge(rev_kl_week[["week_normalized", "revenue_per_kentleg"]],
                                left_on="week", right_on="week_normalized", how="left")
    trend = trend.drop(columns=["week_normalized"])
    trend = trend.merge(ha_week, left_on="week", right_on="week_normalized", how="left")
    trend = trend.drop(columns=["week_normalized"])
    trend = trend.merge(nb_week, left_on="week", right_on="week_normalized", how="left")
    trend = trend.drop(columns=["week_normalized"])
    trend = trend.merge(conc_week_df, on="week", how="left")
    trend = trend.merge(wm_week, left_on="week", right_on="week_normalized", how="left")
    if "week_normalized" in trend.columns:
        trend = trend.drop(columns=["week_normalized"])

    # Week-over-week deltas
    for col in ["vehicle_usage", "billed_usage", "volume_pool_ratio", "otp",
                 "revenue_per_kentleg", "higher_acuity_mix", "nonbillable_ns_rate",
                 "largest_payer_vol", "largest_payer_rev", "profit_margin"]:
        if col in trend.columns:
            trend[f"{col}_wow"] = trend[col].diff()

    _save_csv(trend, REPORTS / "weekly_gate_trend.csv", result)

    # Plot: small-multiples of weekly trend for 6 key gates
    metrics = [
        ("vehicle_usage", "Vehicle utilization", 0.95, "ge"),
        ("billed_usage", "Billed utilization", 1.05, "ge"),
        ("volume_pool_ratio", "Volume pool (SKL/Target)", 1.20, "ge"),
        ("revenue_per_kentleg", "Revenue / Kent-Leg ($)", 70.0, "ge"),
        ("higher_acuity_mix", "Higher-acuity mix", 0.05, "ge"),
        ("nonbillable_ns_rate", "Non-billable NS rate", 0.10, "le"),
    ]
    fig, axes = plt.subplots(2, 3, figsize=(14, 8))
    for ax, (col, title, target, cmp_) in zip(axes.flatten(), metrics):
        if col not in trend.columns:
            ax.axis("off")
            continue
        ax.plot(trend["week"], trend[col], marker="o", color="#00274C")
        ax.axhline(target, ls="--", color="#FFCB05", label=f"target {target}")
        ax.set_title(title)
        ax.tick_params(axis="x", rotation=30)
    fig.suptitle("Weekly trend - Q1 2026")
    _savefig(fig, PLOTS / "weekly_trend_small_multiples.png", result)

    result.summary["weeks_covered"] = len(trend)
    return trend


# ----- section 4: mode mix -----------------------------------------------


def section_mode(result: EdaResult, cv: pd.DataFrame) -> pd.DataFrame:
    completed = cv[cv["order_status"] == "Completed"].copy()
    modes = cv["order_mode"].dropna().unique().tolist()
    rows = []
    total_trips = len(completed)
    total_kl = completed["kent_legs"].sum()
    total_rev = completed["order_price"].sum()
    for mode in modes:
        sub = completed[completed["order_mode"] == mode]
        ns_sub = cv[(cv["order_mode"] == mode) &
                     (cv["reason"].isin(["No Show", "No show", "Same Day Cancel", "Confirmation Cancel"])) &
                     (cv["order_status"] != "Completed")]
        billable_ns = cv[(cv["order_mode"] == mode) & (cv["reason"] == "Billed no show")]
        denom = cv[cv["order_mode"] == mode]
        rows.append({
            "mode": mode,
            "trip_count": len(sub),
            "trip_share": len(sub) / total_trips if total_trips else None,
            "kent_legs": float(sub["kent_legs"].sum()),
            "kent_leg_share": float(sub["kent_legs"].sum() / total_kl) if total_kl else None,
            "revenue": float(sub["order_price"].sum()),
            "revenue_share": float(sub["order_price"].sum() / total_rev) if total_rev else None,
            "avg_revenue_per_trip": float(sub["order_price"].mean()) if len(sub) else None,
            "avg_revenue_per_kentleg": float(sub["order_price"].sum() / sub["kent_legs"].sum()) if sub["kent_legs"].sum() else None,
            "avg_miles": float(sub["order_mileage"].mean()) if len(sub) else None,
            "nonbillable_ns_rate": (len(ns_sub) / len(denom)) if len(denom) else None,
            "billable_ns_rate": (len(billable_ns) / len(denom)) if len(denom) else None,
        })

    # Add SecureCare from securecare_profit_base
    sc = pd.read_csv(PHASE1 / "securecare_profit_base.csv")
    sc_q1 = sc[sc["scope_level"] == "quarter_total"]
    sc_rev = sc_q1["total_revenue"].sum()
    sc_cost = sc_q1["total_cost"].sum()
    rows.append({
        "mode": "SecureCare",
        "trip_count": None,
        "trip_share": None,
        "kent_legs": None,
        "kent_leg_share": None,
        "revenue": float(sc_rev),
        "revenue_share": float(sc_rev / total_rev) if total_rev else None,
        "avg_revenue_per_trip": None,
        "avg_revenue_per_kentleg": None,
        "avg_miles": None,
        "nonbillable_ns_rate": None,
        "billable_ns_rate": None,
        "profit_margin": float((sc_rev - sc_cost) / sc_rev) if sc_rev else None,
        "total_cost": float(sc_cost),
        "note": "From SecureCare Profit sheet; trip log not per-trip",
    })

    df = pd.DataFrame(rows)
    _save_csv(df, REPORTS / "mode_profitability.csv", result)

    # Plot: mix donut + revenue bars
    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    base = df[df["mode"] != "SecureCare"]
    axes[0].pie(base["trip_count"], labels=base["mode"], autopct="%1.1f%%",
                colors=["#00274C", "#FFCB05", "#9aa3ad"])
    axes[0].set_title("Completed-trip mix (Q1)")
    axes[1].bar(df["mode"], df["avg_revenue_per_kentleg"].fillna(0), color="#00274C")
    axes[1].axhline(70.0, ls="--", color="red", label="$70 target")
    axes[1].set_ylabel("Revenue / Kent-Leg")
    axes[1].set_title("Revenue / Kent-Leg by mode")
    axes[1].legend()
    _savefig(fig, PLOTS / "mode_mix.png", result)

    result.summary["modes_analyzed"] = int(len(df))
    return df


# ----- section 5: OTP matrix ---------------------------------------------


def _parse_otp_sheet(ws) -> pd.DataFrame:
    """Parse the OTP sheet into tidy (scope, region, leg, day, otp) rows.

    The sheet is laid out as alternating "Quarter Total" / "Week N" blocks
    where the right half (col offset ~10) contains precomputed OTP
    percentages per location.
    """

    rows = list(ws.iter_rows(values_only=True))
    records: List[Dict[str, object]] = []

    def _leg_from_label(label: str) -> Optional[str]:
        if "A Leg" in label:
            return "A-leg"
        if "B Leg" in label:
            return "B-leg"
        if "Overall" in label:
            return "overall"
        return None

    current_scope: Optional[str] = None
    for ridx, row in enumerate(rows):
        col1 = row[1] if len(row) > 1 else None
        col10 = row[10] if len(row) > 10 else None
        col11 = row[11] if len(row) > 11 else None
        if isinstance(col1, str):
            v = col1.strip()
            if v == "Quarter Total":
                current_scope = "Quarter Total"
            elif v in WEEK_ORDER:
                current_scope = v

        # Right-hand block has Location + metric label + Mon..Fri OTP %
        if isinstance(col10, str) and col10.strip() and isinstance(col11, str):
            location = col10.strip()
            leg = _leg_from_label(col11.strip())
            if not leg:
                continue
            if location not in REGIONS + ["Total", "Detroit"]:
                continue
            mon = _num(row[12]) if len(row) > 12 else None
            tue = _num(row[13]) if len(row) > 13 else None
            wed = _num(row[14]) if len(row) > 14 else None
            thu = _num(row[15]) if len(row) > 15 else None
            fri = _num(row[16]) if len(row) > 16 else None
            tot = _num(row[17]) if len(row) > 17 else None
            for day, val in zip(["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Total"],
                                 [mon, tue, wed, thu, fri, tot]):
                records.append({
                    "scope": current_scope,
                    "region": location,
                    "leg": leg,
                    "day": day,
                    "otp": val,
                })

    return pd.DataFrame(records)


def section_otp(result: EdaResult) -> pd.DataFrame:
    wb = _open_q1()
    otp = _parse_otp_sheet(wb["OTP"])
    wb.close()

    # Drop empty Detroit rows (future market, no Q1 data)
    detroit_has_data = otp[(otp["region"] == "Detroit") & (otp["otp"].notna()) & (otp["otp"] > 0)]
    if detroit_has_data.empty:
        otp = otp[otp["region"] != "Detroit"]

    _save_csv(otp, REPORTS / "otp_matrix.csv", result)

    # Heatmap: A-leg OTP by region x day, Quarter Total
    q1_aleg = otp[(otp["scope"] == "Quarter Total") & (otp["leg"] == "A-leg")]
    pivot_a = q1_aleg.pivot_table(index="region", columns="day", values="otp", aggfunc="mean")
    pivot_a = pivot_a.reindex(index=["Total"] + REGIONS)
    pivot_a = pivot_a.reindex(columns=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Total"])

    q1_bleg = otp[(otp["scope"] == "Quarter Total") & (otp["leg"] == "B-leg")]
    pivot_b = q1_bleg.pivot_table(index="region", columns="day", values="otp", aggfunc="mean")
    pivot_b = pivot_b.reindex(index=["Total"] + REGIONS)
    pivot_b = pivot_b.reindex(columns=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Total"])

    # Use a fallback: if Quarter Total is all zeros (divided-by-zero upstream),
    # average the weekly blocks instead.
    if pivot_a.fillna(0).values.sum() == 0:
        src = otp[(otp["scope"].isin(WEEK_ORDER)) & (otp["leg"] == "A-leg")]
        pivot_a = src.pivot_table(index="region", columns="day", values="otp", aggfunc="mean")
        pivot_a = pivot_a.reindex(index=["Total"] + REGIONS)
        pivot_a = pivot_a.reindex(columns=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Total"])
    if pivot_b.fillna(0).values.sum() == 0:
        src = otp[(otp["scope"].isin(WEEK_ORDER)) & (otp["leg"] == "B-leg")]
        pivot_b = src.pivot_table(index="region", columns="day", values="otp", aggfunc="mean")
        pivot_b = pivot_b.reindex(index=["Total"] + REGIONS)
        pivot_b = pivot_b.reindex(columns=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Total"])

    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    for ax, piv, title in zip(axes, [pivot_a, pivot_b], ["A-leg OTP", "B-leg OTP"]):
        im = ax.imshow(piv.values.astype(float), aspect="auto", cmap="RdYlGn",
                        vmin=0.5, vmax=1.0)
        ax.set_xticks(range(piv.shape[1]))
        ax.set_xticklabels(piv.columns, rotation=30)
        ax.set_yticks(range(piv.shape[0]))
        ax.set_yticklabels(piv.index)
        ax.set_title(f"{title} (Q1, target 90%)")
        for i in range(piv.shape[0]):
            for j in range(piv.shape[1]):
                val = piv.values[i, j]
                if val and not np.isnan(val):
                    ax.text(j, i, f"{val:.0%}", ha="center", va="center", fontsize=8)
        fig.colorbar(im, ax=ax, fraction=0.04)
    _savefig(fig, PLOTS / "otp_heatmap.png", result)

    result.summary["otp_rows"] = int(len(otp))
    return otp


# ----- section 6: contract patterns --------------------------------------


def section_contract_patterns(result: EdaResult, cv: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Payer concentration + cancellation patterns."""

    completed = cv[cv["order_status"] == "Completed"].copy()
    tot_vol = completed["kent_legs"].sum()
    tot_rev = completed["order_price"].sum()

    by_payer = completed.groupby("payer_id").agg(
        kent_legs=("kent_legs", "sum"),
        revenue=("order_price", "sum"),
        trips=("order_price", "size"),
    ).reset_index()
    by_payer["vol_share"] = by_payer["kent_legs"] / tot_vol
    by_payer["rev_share"] = by_payer["revenue"] / tot_rev
    by_payer["over_20pct_vol"] = by_payer["vol_share"] > 0.20
    by_payer["over_20pct_rev"] = by_payer["rev_share"] > 0.20
    by_payer["near_cap"] = (by_payer["vol_share"] > 0.17) | (by_payer["rev_share"] > 0.17)
    by_payer = by_payer.sort_values("rev_share", ascending=False)
    _save_csv(by_payer, REPORTS / "payer_concentration.csv", result)

    # Cancellation patterns: row per (payer, status, reason)
    canc = cv[cv["order_status"] != "Completed"].copy()
    patterns = canc.groupby(["order_status", "reason", "payer_id", "order_mode", "day"]).size().reset_index(name="count")
    patterns = patterns.sort_values("count", ascending=False)
    _save_csv(patterns, REPORTS / "cancellation_patterns.csv", result)

    # Plot 1: pareto of payer revenue share (top 15)
    fig, ax = plt.subplots(figsize=(12, 5))
    top = by_payer.head(15)
    ax.bar(top["payer_id"], top["rev_share"], color="#00274C")
    ax.axhline(0.20, color="red", ls="--", label="20% cap")
    ax.set_ylabel("Revenue share")
    ax.set_title("Top 15 payers by revenue share (Gate 8: ≤20%)")
    ax.tick_params(axis="x", rotation=60)
    ax.legend()
    _savefig(fig, PLOTS / "payer_pareto.png", result)

    # Plot 2: cancellation reasons stacked by mode
    by_reason_mode = patterns.groupby(["reason", "order_mode"])["count"].sum().reset_index()
    piv = by_reason_mode.pivot_table(index="reason", columns="order_mode", values="count", aggfunc="sum").fillna(0)
    fig, ax = plt.subplots(figsize=(10, 5))
    piv.plot(kind="bar", stacked=True, ax=ax, color=["#00274C", "#FFCB05", "#9aa3ad"])
    ax.set_title("Cancellation / no-show reasons by mode (Q1)")
    ax.set_ylabel("Trip count")
    ax.tick_params(axis="x", rotation=30)
    _savefig(fig, PLOTS / "cancellation_by_reason.png", result)

    result.summary["payers_total"] = int(len(by_payer))
    result.summary["payers_over_20pct_vol"] = int(by_payer["over_20pct_vol"].sum())
    result.summary["payers_over_20pct_rev"] = int(by_payer["over_20pct_rev"].sum())
    return by_payer, patterns


# ----- section 7: hourly demand / idle ------------------------------------


def _parse_heat_map(ws) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Return (total_rides_hourly, total_completed_hourly) dataframes."""

    rows = list(ws.iter_rows(values_only=True))
    # Block 1: "Total Rides Q4" → header at row 2, data rows 3-9
    # Block 2: "Total Rides Completed" → header at row 11, data rows 12-18
    # Structure is stable: find header rows by matching "Day" in col 1 after
    # a "Total Rides" banner in col 1 one row above.
    def _block_to_df(header_idx: int) -> pd.DataFrame:
        hdr = rows[header_idx]
        hours = [str(hdr[c]) for c in range(2, 26)]  # 00:00..22:00 + will call + Total
        out = []
        for r in rows[header_idx + 1: header_idx + 8]:
            label = str(r[1]) if r[1] else ""
            if label in DAY_ORDER + ["Total"]:
                for c, hour in zip(range(2, 26), hours):
                    out.append({
                        "day": label,
                        "hour": hour,
                        "value": _num(r[c]),
                    })
        return pd.DataFrame(out)

    requested = _block_to_df(2)
    completed = _block_to_df(11)
    return requested, completed


def section_heatmap(result: EdaResult) -> pd.DataFrame:
    wb = _open_q1()
    requested, completed = _parse_heat_map(wb["Heat Map"])
    wb.close()

    requested["metric"] = "requested"
    completed["metric"] = "completed"
    both = pd.concat([requested, completed], ignore_index=True)

    # Idle windows: for each (day, hour), what's the completed count?
    # A window is idle if completed == 0 during weekday business hours.
    idle = completed[(~completed["day"].isin(["Total"])) &
                     (completed["value"].fillna(0) < 1)]
    idle = idle[~idle["hour"].isin(["Total"])]
    idle["hour_of_day"] = idle["hour"].str.split(":").str[0]
    idle_business = idle[idle["hour_of_day"].astype(str).isin(["06", "07", "08", "09", "10",
                                                                  "11", "12", "13", "14", "15",
                                                                  "16", "17", "18"])]
    both["is_idle_business"] = False
    both.loc[(both["metric"] == "completed") &
              (both["day"].isin(DAY_ORDER)) &
              (both["hour"].str.startswith(("06", "07", "08", "09", "1", "18"))) &
              (both["value"].fillna(0) < 1), "is_idle_business"] = True

    _save_csv(both, REPORTS / "hourly_demand_idle.csv", result)

    # Plot heatmap of completed trips by day×hour
    piv = completed[(~completed["day"].isin(["Total"])) &
                     (~completed["hour"].isin(["Total", "will call"]))].pivot_table(
        index="day", columns="hour", values="value", aggfunc="first"
    ).reindex(index=DAY_ORDER)
    hours_sorted = sorted([h for h in piv.columns if isinstance(h, str) and ":" in h])
    piv = piv[hours_sorted]

    fig, ax = plt.subplots(figsize=(14, 5))
    im = ax.imshow(piv.values.astype(float), aspect="auto", cmap="YlOrRd")
    ax.set_xticks(range(len(hours_sorted)))
    ax.set_xticklabels([h[:5] for h in hours_sorted], rotation=45)
    ax.set_yticks(range(len(DAY_ORDER)))
    ax.set_yticklabels(DAY_ORDER)
    ax.set_title("Completed trips by hour × day (Q4 baseline)")
    fig.colorbar(im, ax=ax)
    _savefig(fig, PLOTS / "hourly_heatmap.png", result)

    result.summary["idle_windows_business"] = int(len(idle_business))
    return both


# ----- section 8: revenue by payer ---------------------------------------


def section_revenue_payer(result: EdaResult, cv: pd.DataFrame) -> pd.DataFrame:
    completed = cv[cv["order_status"] == "Completed"].copy()

    by_payer_mode = completed.groupby(["payer_id", "order_mode"]).agg(
        revenue=("order_price", "sum"),
        kent_legs=("kent_legs", "sum"),
        trips=("order_price", "size"),
    ).reset_index()
    by_payer_mode["revenue_per_kentleg"] = by_payer_mode["revenue"] / by_payer_mode["kent_legs"]

    by_payer = completed.groupby("payer_id").agg(
        revenue=("order_price", "sum"),
        kent_legs=("kent_legs", "sum"),
        trips=("order_price", "size"),
    ).reset_index()
    by_payer["revenue_per_kentleg"] = by_payer["revenue"] / by_payer["kent_legs"]
    by_payer["lift_vs_70"] = by_payer["revenue_per_kentleg"] - 70.0

    _save_csv(by_payer_mode, REPORTS / "payer_rev_per_kentleg_by_mode.csv", result)
    _save_csv(by_payer, REPORTS / "payer_rev_per_kentleg.csv", result)

    # Waterfall plot: payers ranked by lift (top 20 by |lift| x volume)
    top = by_payer.sort_values("kent_legs", ascending=False).head(20).copy()
    top = top.sort_values("lift_vs_70")
    fig, ax = plt.subplots(figsize=(12, 6))
    colors = ["#A80000" if v < 0 else "#1f8a1f" for v in top["lift_vs_70"]]
    ax.barh(top["payer_id"], top["lift_vs_70"], color=colors)
    ax.axvline(0, color="black", lw=0.8)
    ax.set_xlabel("Revenue / Kent-Leg minus $70 target")
    ax.set_title("Top 20 payers by volume - lift vs $70 Kent-Leg target")
    _savefig(fig, PLOTS / "payer_rev_waterfall.png", result)

    overall = completed["order_price"].sum() / completed["kent_legs"].sum() if completed["kent_legs"].sum() else None
    result.summary["fleet_rev_per_kentleg_q1"] = float(overall) if overall else None
    return by_payer


# ----- section 9: cost & margin ------------------------------------------


def section_cost_margin(result: EdaResult) -> Tuple[pd.DataFrame, pd.DataFrame]:
    wm = pd.read_csv(PHASE1 / "weekly_margin_base.csv")
    sc = pd.read_csv(PHASE1 / "securecare_profit_base.csv")
    vd = pd.read_csv(PHASE1 / "vehicle_day_base.csv")

    # Fleet weekly totals (daily rows in weekly_margin_base are scope_level="week";
    # roll them up to one row per (stream, week).)
    def _agg(df: pd.DataFrame, stream: str) -> pd.DataFrame:
        g = df.groupby("week_normalized").agg(
            total_revenue=("total_revenue", "sum"),
            total_cost=("total_cost", "sum"),
        ).reset_index().rename(columns={"week_normalized": "week"})
        g["profit_margin"] = (g["total_revenue"] - g["total_cost"]) / g["total_revenue"]
        g["stream"] = stream
        return g[["week", "stream", "total_revenue", "total_cost", "profit_margin"]]

    wm_week = _agg(wm[wm["scope_level"] == "week"], "Fleet")
    sc_week = _agg(sc[sc["scope_level"] == "week"], "SecureCare")

    cost_trend = pd.concat([wm_week, sc_week], ignore_index=True)
    # Preserve chronological order
    cost_trend["_ord"] = cost_trend["week"].map({w: i for i, w in enumerate(WEEK_ORDER)})
    cost_trend = cost_trend.sort_values(["stream", "_ord"]).drop(columns=["_ord"]).reset_index(drop=True)
    _save_csv(cost_trend, REPORTS / "cost_margin_trend.csv", result)

    # Regional cost estimate: apportion total cost by regional road hours
    total_cost_q1 = wm[wm["scope_level"] == "quarter_total"]["total_cost"].sum()
    total_road_hours = vd["road_time"].sum()

    # No vehicle→region mapping available; allocate by published fleet size
    # (21 GR, 4 L, 5 BC) as a proxy for road-hour share.
    fleet_sizes = {"Grand Rapids": 21, "Lansing": 4, "Battle Creek": 5}
    fleet_total = sum(fleet_sizes.values())
    rows = []
    for region, n in fleet_sizes.items():
        share = n / fleet_total
        region_cost = total_cost_q1 * share
        region_road_hours = total_road_hours * share
        rows.append({
            "region": region,
            "vehicle_count": n,
            "cost_share_assumed": share,
            "estimated_cost": float(region_cost),
            "estimated_road_hours": float(region_road_hours),
            "estimated_cost_per_road_hour": float(region_cost / region_road_hours) if region_road_hours else None,
            "note": "Estimate; RYW has not broken cost down per region yet",
        })
    regional_cost = pd.DataFrame(rows)
    _save_csv(regional_cost, REPORTS / "regional_cost_estimate.csv", result)

    # Plot: margin trend line (fleet + SecureCare)
    fig, ax = plt.subplots(figsize=(10, 5))
    for stream, group in cost_trend.groupby("stream"):
        agg = group.groupby("week")["profit_margin"].mean().reindex(WEEK_ORDER)
        ax.plot(agg.index, agg.values, marker="o", label=stream)
    ax.axhline(0.25, ls="--", color="red", label="25% target")
    ax.set_title("Weekly profit margin - Fleet vs SecureCare")
    ax.legend()
    _savefig(fig, PLOTS / "margin_trend.png", result)

    result.summary["total_cost_q1"] = float(total_cost_q1)
    result.summary["total_road_hours_q1"] = float(total_road_hours)
    return cost_trend, regional_cost


# ----- section 10: cross-views -------------------------------------------


def section_crossview(result: EdaResult, cv: pd.DataFrame) -> pd.DataFrame:
    """Gate x Fleet x Week master table."""

    wb = _open_q1()
    regional = _extract_regional(wb["Regional Performance"])
    wb.close()

    weekly_totals = regional[
        (regional["label"] == "Total") &
        (regional["scope"].isin(WEEK_ORDER))
    ].copy()

    records = []
    for _, r in weekly_totals.iterrows():
        records.append({
            "region": r["region"],
            "week": r["scope"],
            "vehicle_utilization": r["vehicle_usage"],
            "billed_utilization": r["billed_usage"],
            "total_pool": r["total_pool"],
            "otp": r["otp"],
            "total_rides": r["total_rides"],
            "kent_legs": r["kent_legs"],
        })

    # Revenue/KL + higher-acuity mix per (region, week)
    completed = cv[cv["order_status"] == "Completed"].copy()
    completed["pu_zone_norm"] = completed["pu_zone"].replace({
        "GrandRapids": "Grand Rapids",
        "Battle Creek/Kalamazoo": "Battle Creek",
    })
    for rec in records:
        sub = completed[(completed["pu_zone_norm"] == rec["region"]) &
                         (completed["week_normalized"] == rec["week"])]
        if len(sub):
            rec["revenue_per_kentleg"] = (sub["order_price"].sum() / sub["kent_legs"].sum()) if sub["kent_legs"].sum() else None
            rec["higher_acuity_mix"] = (sub["order_mode"] == "Stretcher").sum() / len(sub)
        else:
            rec["revenue_per_kentleg"] = None
            rec["higher_acuity_mix"] = None

    df = pd.DataFrame(records)
    _save_csv(df, REPORTS / "gate_fleet_week_crossview.csv", result)

    # Plot: overall heatmap of gate health (region rows × week columns)
    metric = "vehicle_utilization"
    piv = df.pivot_table(index="region", columns="week", values=metric, aggfunc="mean")
    piv = piv.reindex(index=REGIONS)[WEEK_ORDER]
    fig, ax = plt.subplots(figsize=(9, 4))
    im = ax.imshow(piv.values.astype(float), aspect="auto", cmap="RdYlGn", vmin=0.5, vmax=1.2)
    ax.set_xticks(range(len(WEEK_ORDER)))
    ax.set_xticklabels(WEEK_ORDER)
    ax.set_yticks(range(len(REGIONS)))
    ax.set_yticklabels(REGIONS)
    for i in range(piv.shape[0]):
        for j in range(piv.shape[1]):
            val = piv.values[i, j]
            if val and not np.isnan(val):
                ax.text(j, i, f"{val:.0%}", ha="center", va="center", fontsize=9)
    ax.set_title("Vehicle utilization - region × week")
    fig.colorbar(im, ax=ax)
    _savefig(fig, PLOTS / "gate_fleet_week_heatmap.png", result)

    return df


# ----- entry point --------------------------------------------------------


def run() -> EdaResult:
    _ensure_dirs()
    result = EdaResult()

    print(f"[operational_eda] writing reports to {REPORTS}")
    print(f"[operational_eda] writing plots to {PLOTS}")

    print("[1/10] Coverage...")
    section_coverage(result)

    # Shared: load the completed/cancel trip log
    cv = pd.read_csv(PHASE1 / "contract_volume_base.csv")

    print("[2/10] Fleetwise...")
    section_fleetwise(result, cv)

    print("[3/10] Weekly trend...")
    section_weekly(result, cv)

    print("[4/10] Mode mix...")
    section_mode(result, cv)

    print("[5/10] OTP...")
    section_otp(result)

    print("[6/10] Contract patterns...")
    section_contract_patterns(result, cv)

    print("[7/10] Hourly heatmap...")
    section_heatmap(result)

    print("[8/10] Revenue by payer...")
    section_revenue_payer(result, cv)

    print("[9/10] Cost & margin...")
    section_cost_margin(result)

    print("[10/10] Cross-views...")
    section_crossview(result, cv)

    # Manifest
    manifest = {
        "csv_artifacts": [str(p) for p in result.csvs],
        "plot_artifacts": [str(p) for p in result.plots],
        "summary": result.summary,
    }
    manifest_path = REPORTS / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, default=str))
    result.add_csv(manifest_path)  # treat as an artifact

    print(f"[operational_eda] wrote {len(result.csvs)} CSVs and {len(result.plots)} plots")
    return result


if __name__ == "__main__":
    try:
        run()
    except Exception as e:  # pragma: no cover
        print(f"operational_eda failed: {e}", file=sys.stderr)
        raise
